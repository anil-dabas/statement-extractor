from typing import List, Dict
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from core.transaction import Transaction
from core.models import NATURE_OPTIONS


class ExcelExporter:
    """Exports transactions to Excel in the Accounting Queries format."""

    def __init__(self):
        self.header_font = Font(bold=True)
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def export(
        self,
        transactions: List[Transaction],
        output_path: str,
        customer_name: str = "",
    ) -> str:
        """
        Export transactions to Excel file matching Accounting Queries format.

        Args:
            transactions: List of Transaction objects
            output_path: Path for the output Excel file
            customer_name: Customer name for the header

        Returns:
            Path to the created Excel file
        """
        wb = Workbook()

        # Create sheets
        bank_in_sheet = wb.active
        bank_in_sheet.title = "Bank In"

        bank_out_sheet = wb.create_sheet("Bank Out")
        sheet3 = wb.create_sheet("Sheet3")

        # Separate transactions by type
        bank_in = [t for t in transactions if t.transaction_type == "in"]
        bank_out = [t for t in transactions if t.transaction_type == "out"]

        # Sort by currency then by date
        bank_in.sort(key=lambda t: (t.currency, t.date))
        bank_out.sort(key=lambda t: (t.currency, t.date))

        # Determine customer name
        final_customer_name = customer_name
        if not final_customer_name and transactions:
            # Use the first transaction's customer name
            final_customer_name = transactions[0].customer_name

        # Write sheets
        self._write_transaction_sheet(bank_in_sheet, bank_in, "Bank In", final_customer_name)
        self._write_transaction_sheet(bank_out_sheet, bank_out, "Bank Out", final_customer_name)
        self._write_nature_sheet(sheet3)

        # Save workbook
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return str(output_path)

    def _write_transaction_sheet(
        self,
        sheet,
        transactions: List[Transaction],
        sheet_title: str,
        customer_name: str
    ):
        """Write transactions to a sheet in the Accounting Queries format."""

        # Row 1: Customer Name
        sheet.cell(row=1, column=1, value=customer_name).font = self.header_font

        # Row 3: Sheet title (Bank In / Bank Out)
        sheet.cell(row=3, column=1, value=sheet_title).font = self.header_font

        # Group transactions by currency
        currency_groups: Dict[str, List[Transaction]] = defaultdict(list)
        for t in transactions:
            currency_groups[t.currency].append(t)

        # Define currency order (HKD first, then others alphabetically)
        currency_order = ["HKD"] + sorted([c for c in currency_groups.keys() if c != "HKD"])

        current_row = 5
        data_start_row = None

        for currency in currency_order:
            if currency not in currency_groups:
                continue

            currency_transactions = currency_groups[currency]

            # Currency section header
            account_type = f"{currency} Saving Account"
            sheet.cell(row=current_row, column=1, value=account_type).font = self.header_font

            # Column headers in same row
            sheet.cell(row=current_row, column=4, value=currency if currency == "HKD" else "HKD")
            sheet.cell(row=current_row, column=6, value="")  # Exchange rate column
            sheet.cell(row=current_row, column=7, value="Description")
            sheet.cell(row=current_row, column=9, value="Nature")
            sheet.cell(row=current_row, column=11, value="Remark")

            current_row += 1
            if data_start_row is None:
                data_start_row = current_row

            # Write transactions for this currency
            for transaction in currency_transactions:
                # Column B: Date
                sheet.cell(row=current_row, column=2, value=transaction.date)
                sheet.cell(row=current_row, column=2).number_format = "yyyy-mm-dd"

                # Column D: Amount (convert to HKD if needed, or show original)
                if currency == "HKD":
                    sheet.cell(row=current_row, column=4, value=float(transaction.amount))
                else:
                    # For foreign currency, amount stays as original
                    # User can add exchange rate formula
                    sheet.cell(row=current_row, column=4, value=float(transaction.amount))

                # Column F: Exchange rate comment (for foreign currency)
                if currency != "HKD":
                    exchange_note = f"{currency}{transaction.amount}"
                    if transaction.exchange_rate:
                        exchange_note += f" @{transaction.exchange_rate}"
                    sheet.cell(row=current_row, column=6, value=exchange_note)

                # Column G: Description
                sheet.cell(row=current_row, column=7, value=transaction.description)

                # Column I: Nature (empty for user to fill via dropdown)
                sheet.cell(row=current_row, column=9, value=transaction.nature)

                # Column K: Remark
                sheet.cell(row=current_row, column=11, value=transaction.remark)

                current_row += 1

            # Add empty row between currency sections
            current_row += 1

        # Set column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 5
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 5
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 35
        sheet.column_dimensions['H'].width = 5
        sheet.column_dimensions['I'].width = 25
        sheet.column_dimensions['J'].width = 5
        sheet.column_dimensions['K'].width = 25

        # Add data validation for Nature column if there are transactions
        if data_start_row and current_row > data_start_row:
            self._add_nature_validation(sheet, data_start_row, current_row - 1)

    def _write_nature_sheet(self, sheet):
        """Write nature options to Sheet3."""
        # First row: "Please Select"
        sheet.cell(row=1, column=1, value="Please Select")

        # Nature options
        for row_idx, option in enumerate(NATURE_OPTIONS, 2):
            sheet.cell(row=row_idx, column=1, value=option)

        # Adjust column width
        sheet.column_dimensions["A"].width = 30

    def _add_nature_validation(self, sheet, start_row: int, end_row: int):
        """Add dropdown validation for Nature column (column I)."""
        # Create formula reference to Sheet3 nature options
        nature_range = f"Sheet3!$A$1:$A${len(NATURE_OPTIONS) + 1}"

        dv = DataValidation(
            type="list",
            formula1=nature_range,
            allow_blank=True,
            showDropDown=False,  # False means show dropdown
        )
        dv.error = "Please select a valid Nature option"
        dv.errorTitle = "Invalid Nature"
        dv.prompt = "Select nature from dropdown"
        dv.promptTitle = "Nature"

        # Apply to Nature column (column I)
        dv.add(f"I{start_row}:I{end_row}")
        sheet.add_data_validation(dv)


def create_excel_from_transactions(
    transactions: List[Transaction],
    output_path: str,
    customer_name: str = "",
) -> str:
    """
    Convenience function to create Excel file from transactions.

    Args:
        transactions: List of Transaction objects
        output_path: Path for the output Excel file
        customer_name: Default customer name

    Returns:
        Path to the created Excel file
    """
    exporter = ExcelExporter()
    return exporter.export(transactions, output_path, customer_name)
